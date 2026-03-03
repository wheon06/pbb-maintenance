from __future__ import annotations

from io import BytesIO
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter

import config
from models import FlightItem, FlightType
from services import TaggedFlight
from utils import format_date, format_time

THIN_BORDER = Border(
    left=Side(style="thin", color="CCCCCC"),
    right=Side(style="thin", color="CCCCCC"),
    top=Side(style="thin", color="CCCCCC"),
    bottom=Side(style="thin", color="CCCCCC"),
)


def _resolve_cell_value(item: FlightItem, field: str, flight_type: FlightType) -> str:
    if field == "_date":
        return format_date(item.scheduled_datetime)

    elif field == "_flight_type":
        return flight_type.value

    elif field in ("scheduled_datetime", "actual_datetime"):
        return format_time(getattr(item, field, ""))

    elif field == "_departure_airport":
        return item.airport_name or "-" if flight_type is FlightType.ARRIVAL else "-"

    elif field == "_arrival_airport":
        return item.airport_name or "-" if flight_type is FlightType.DEPARTURE else "-"

    else:
        return getattr(item, field, "") or "-"


def write_excel_sheet(workbook: Workbook, sheet_name: str, all_items: list[TaggedFlight]):
    worksheet = workbook.create_sheet(title=sheet_name)

    for column_index, (column_name, _) in enumerate(config.EXCEL_COLUMNS, 1):
        cell = worksheet.cell(row=1, column=column_index, value=column_name)
        cell.fill = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
        cell.font = Font(name="맑은 고딕", bold=True, color="FFFFFF", size=10)
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = THIN_BORDER

    master_items = [tf for tf in all_items if tf.item.is_master]
    master_items.sort(key=lambda tf: tf.item.scheduled_datetime)

    for row_index, tagged in enumerate(master_items, 2):
        for column_index, (_, field) in enumerate(config.EXCEL_COLUMNS, 1):
            value = _resolve_cell_value(tagged.item, field, tagged.flight_type)
            cell = worksheet.cell(row=row_index, column=column_index, value=value)
            cell.font = Font(name="맑은 고딕", size=10)
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.border = THIN_BORDER

    for column_index in range(1, len(config.EXCEL_COLUMNS) + 1):
        max_length = 0
        for row in worksheet.iter_rows(min_col=column_index, max_col=column_index, values_only=True):
            cell_value = str(row[0]) if row[0] else ""
            length = sum(2 if ord(character) > 127 else 1 for character in cell_value)
            max_length = max(max_length, length)
        worksheet.column_dimensions[get_column_letter(column_index)].width = max(max_length + 3, 10)

    worksheet.auto_filter.ref = worksheet.dimensions


def create_excel_file(terminal_items: dict[str, list[TaggedFlight]]) -> Workbook:
    workbook = Workbook()
    workbook.remove(workbook.active)

    for terminal in config.TERMINALS:
        items = terminal_items.get(terminal.terminal_id, [])
        write_excel_sheet(workbook, terminal.name, items)

    return workbook


def file_to_bytes_io(workbook: Workbook) -> BytesIO:
    output = BytesIO()
    workbook.save(output)
    output.seek(0)
    return output
